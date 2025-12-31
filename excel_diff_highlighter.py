"""
Excelファイルの差分比較ツール
新旧バージョンのExcelファイルを比較し、差分がある文字のみを青色でハイライトします。
"""

import openpyxl
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
import difflib
import re
import time
from pathlib import Path
from typing import Tuple, List, Optional, Dict
from datetime import datetime
import html
import json

# 設定
MAX_CELL_VALUE_LENGTH = 100  # 差分サマリーシートに表示するセル値の最大文字数
DEFAULT_HIGHLIGHT_COLOR = '000000FF'  # デフォルトのハイライト色（青、aRGB形式）

# ファイル処理設定
OUTPUT_FILE_SUFFIX = "_差分ハイライト"  # 出力ファイル名のサフィックス
TEMP_FILE_PREFIX = '~$'  # 一時ファイルのプレフィックス
PROGRESS_DISPLAY_INTERVAL = 10  # 進捗表示間隔（%）

# UI設定
SEPARATOR_LENGTH = 60  # 区切り線の長さ

# ハイライト色マップ（aRGB形式）
COLOR_MAP = {
    '1': ('000000FF', '青'),
    '2': ('0000FF00', '緑'),
    '3': ('00FF8C00', 'オレンジ'),
    '4': ('00800080', '紫'),
    '5': ('00FF69B4', 'ピンク'),
    '6': ('00FF0000', '赤')
}

# 差分サマリーシート設定
SUMMARY_SHEET_NAME = "差分サマリー"  # 差分サマリーシート名
SUMMARY_HEADER_COLOR = 'D3D3D3'  # ヘッダー背景色（ライトグレー）
SUMMARY_HEADER_FONT_SIZE = 11  # ヘッダーフォントサイズ
SUMMARY_COL_WIDTH_NO = 8       # No.列の幅
SUMMARY_COL_WIDTH_SHEET = 25   # シート名列の幅
SUMMARY_COL_WIDTH_CELL = 10    # セル列の幅
SUMMARY_COL_WIDTH_VALUE = 40   # 旧値/新値列の幅

# HTMLレポート設定
HTML_REPORT_SUFFIX = "_差分レポート"  # HTMLレポートファイル名のサフィックス


def find_file_by_pattern(directory: str, pattern: str) -> List[Path]:
    """
    ファイル名のパターンに基づいてファイルを検索
    例: 「V5_帳票スケッチ_帳票部品No.107_転出名簿4」で検索
    """
    directory_path = Path(directory)
    if not directory_path.exists():
        print(f"エラー: ディレクトリが見つかりません: {directory}")
        return []

    # パターンをエスケープして正規表現として使用
    escaped_pattern = re.escape(pattern)
    # バージョン番号部分を柔軟にマッチ
    regex_pattern = escaped_pattern + r'.*\.xlsx?$'

    matching_files = []
    for file in directory_path.glob('*.xlsx'):
        if re.search(regex_pattern, file.name, re.IGNORECASE):
            matching_files.append(file)

    return matching_files


def extract_version_number(filename: str) -> float:
    """
    ファイル名からバージョン番号を抽出
    例: "v2.06" -> 2.06
    """
    match = re.search(r'[vV](\d+\.\d+)', filename)
    if match:
        return float(match.group(1))
    return 0.0


def extract_base_filename(filename: str) -> str:
    """
    ファイル名からバージョン番号、コピー表記、拡張子を除いた基本名を抽出
    例: "V5_帳票スケッチ_帳票部品No.107_転出名簿4_v2.06.xlsx" -> "V5_帳票スケッチ_帳票部品No.107_転出名簿4"
    例: "V5_帳票スケッチ_帳票部品No.105_転出名簿2_v2.09 のコピー.xlsx" -> "V5_帳票スケッチ_帳票部品No.105_転出名簿2"
    例: "【サイト管理】検索条件・導入元紐づけ整理 のコピー.xlsx" -> "【サイト管理】検索条件・導入元紐づけ整理"
    例: "ファイル名 (1).xlsx" -> "ファイル名"
    """
    # 拡張子を除去
    name_without_ext = Path(filename).stem

    # バージョン番号部分とその後の文字列（のコピー、など）を除去
    # [_\s]*: アンダースコアまたは空白文字（複数可）
    # [vV]\d+\.\d+: バージョン番号（v2.06など）
    # .*$: バージョン番号以降のすべての文字（" のコピー"など）
    base_name = re.sub(r'[_\s]*[vV]\d+\.\d+.*$', '', name_without_ext)

    # バージョン番号がない場合でも「のコピー」「(1)」などを除去
    # \s*: 空白文字（複数可）
    # (のコピー|\(\d+\)|copy): 「のコピー」「(数字)」「copy」などのパターン
    base_name = re.sub(r'\s*(のコピー|\(\d+\)|copy|\s-\s*コピー).*$', '', base_name, flags=re.IGNORECASE)

    return base_name.strip()


def find_matching_file_pairs(old_directory: str, new_directory: str) -> Tuple[List[Tuple[str, str, str]], List[str], List[str]]:
    """
    新旧ディレクトリから対応するファイルペアを検索
    戻り値: (pairs, unmatched_old_files, unmatched_new_files)
        pairs: [(base_name, old_file_path, new_file_path), ...]
        unmatched_old_files: 旧フォルダにのみ存在するファイル名のリスト
        unmatched_new_files: 新フォルダにのみ存在するファイル名のリスト
    """
    old_dir = Path(old_directory)
    new_dir = Path(new_directory)

    if not old_dir.exists():
        print(f"エラー: 旧ディレクトリが見つかりません: {old_directory}")
        return ([], [], [])

    if not new_dir.exists():
        print(f"エラー: 新ディレクトリが見つかりません: {new_directory}")
        return ([], [], [])

    # 旧ディレクトリのファイルを基本名でグループ化
    old_files = {}
    for file in old_dir.glob('*.xlsx'):
        if file.name.startswith(TEMP_FILE_PREFIX):  # 一時ファイルをスキップ
            continue
        base_name = extract_base_filename(file.name)
        if base_name not in old_files:
            old_files[base_name] = []
        old_files[base_name].append(file)

    # 新ディレクトリのファイルを基本名でグループ化
    new_files = {}
    for file in new_dir.glob('*.xlsx'):
        if file.name.startswith(TEMP_FILE_PREFIX):  # 一時ファイルをスキップ
            continue
        base_name = extract_base_filename(file.name)
        if base_name not in new_files:
            new_files[base_name] = []
        new_files[base_name].append(file)

    # マッチングするペアを検索
    pairs = []
    matched_bases = set()

    for base_name in old_files:
        if base_name in new_files:
            # 各グループ内でファイルを選択
            # バージョン番号がある場合は最新を選択、ない場合は最初のファイルを選択
            old_versions = [(f, extract_version_number(f.name)) for f in old_files[base_name]]
            new_versions = [(f, extract_version_number(f.name)) for f in new_files[base_name]]

            # バージョン番号が存在する場合（0.0より大きい）は最新を選択
            if any(v > 0 for _, v in old_versions):
                old_file = max(old_files[base_name], key=lambda f: extract_version_number(f.name))
            else:
                # バージョン番号がない場合は最初のファイル
                old_file = old_files[base_name][0]

            if any(v > 0 for _, v in new_versions):
                new_file = max(new_files[base_name], key=lambda f: extract_version_number(f.name))
            else:
                # バージョン番号がない場合は最初のファイル
                new_file = new_files[base_name][0]

            # 同じファイルでないことを確認（パスが異なる場合は処理）
            if str(old_file) != str(new_file):
                pairs.append((base_name, str(old_file), str(new_file)))
                matched_bases.add(base_name)

    # マッチングしないファイルを報告
    unmatched_old = set(old_files.keys()) - matched_bases
    unmatched_new = set(new_files.keys()) - matched_bases

    # マッチングしなかったファイル名をリストに格納
    unmatched_old_files = []
    unmatched_new_files = []

    if unmatched_old:
        print(f"\n⚠ 旧フォルダにのみ存在するファイル（新バージョンなし）:")
        for base_name in sorted(unmatched_old):
            for file in old_files[base_name]:
                print(f"  - {file.name}")
                unmatched_old_files.append(file.name)

    if unmatched_new:
        print(f"\n⚠ 新フォルダにのみ存在するファイル（旧バージョンなし）:")
        for base_name in sorted(unmatched_new):
            for file in new_files[base_name]:
                print(f"  - {file.name}")
                unmatched_new_files.append(file.name)

    return pairs, unmatched_old_files, unmatched_new_files


def find_old_and_new_versions(directory: str, base_filename: str) -> Tuple[Optional[str], Optional[str]]:
    """
    指定されたディレクトリから新旧のバージョンファイルを検索
    """
    files = find_file_by_pattern(directory, base_filename)

    if not files or len(files) < 2:
        print(f"エラー: {base_filename} に一致するファイルが2つ以上見つかりません")
        if files:
            print(f"見つかったファイル: {[f.name for f in files]}")
        return None, None

    # バージョン番号でソート
    sorted_files = sorted(files, key=lambda f: extract_version_number(f.name))

    old_file = str(sorted_files[-2])  # 2番目に新しいファイル（古いバージョン）
    new_file = str(sorted_files[-1])  # 最新ファイル（新しいバージョン）

    print(f"古いバージョン: {Path(old_file).name}")
    print(f"新しいバージョン: {Path(new_file).name}")

    return old_file, new_file


def get_cell_value_as_string(cell) -> str:
    """
    セルの値を文字列として取得

    Args:
        cell: 対象セル
    """
    if cell.value is None:
        return ""
    return str(cell.value)


def find_char_differences(old_text: str, new_text: str) -> Tuple[List[Tuple[int, int]], str]:
    """
    2つのテキスト間の文字レベルの差分を検出
    戻り値: ([(start_index, end_index), ...], diff_type)
        diff_type: 'insert'(追加), 'delete'(削除), 'replace'(変更), 'equal'(同一)
    """
    if old_text == new_text:
        return [], 'equal'

    # 文字レベルでの差分を検出
    matcher = difflib.SequenceMatcher(None, old_text, new_text)
    differences = []
    diff_types = set()

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == 'insert':
            # 追加された部分
            differences.append((j1, j2))
            diff_types.add('insert')
        elif tag == 'delete':
            # 削除された部分（新テキストには存在しない）
            diff_types.add('delete')
        elif tag == 'replace':
            # 置き換えられた部分
            differences.append((j1, j2))
            diff_types.add('replace')

    # 差分タイプを決定（優先順位: replace > insert > delete）
    if 'replace' in diff_types:
        return differences, 'replace'
    elif 'insert' in diff_types:
        return differences, 'insert'
    elif 'delete' in diff_types:
        return differences, 'delete'
    else:
        return differences, 'equal'


def apply_blue_color_to_differences(cell, old_text: str, new_text: str, highlight_color: str = DEFAULT_HIGHLIGHT_COLOR) -> str:
    """
    差分がある文字のみを指定色にする

    Args:
        cell: 対象セル
        old_text: 旧テキスト
        new_text: 新テキスト
        highlight_color: ハイライト色（aRGB形式の16進数）
    
    Returns:
        diff_type: 差分タイプ ('insert', 'delete', 'replace', 'equal')
    """
    differences, diff_type = find_char_differences(old_text, new_text)

    if not differences:
        return diff_type

    # 元のセルのフォント情報を取得
    original_font = cell.font

    # RichTextオブジェクトを作成
    rich_text_parts = []
    current_pos = 0

    # InlineFontのパラメータを準備（サポートされているもののみ）
    normal_font_kwargs = {}
    blue_font_kwargs = {}

    # フォントサイズ
    if original_font.size:
        normal_font_kwargs['sz'] = original_font.size
        blue_font_kwargs['sz'] = original_font.size

    # フォント名
    if original_font.name:
        normal_font_kwargs['rFont'] = original_font.name
        blue_font_kwargs['rFont'] = original_font.name

    # 元の色（aRGB形式の8文字16進数に変換）
    if original_font.color and original_font.color.rgb:
        try:
            color_value = str(original_font.color.rgb)
            # 16進数カラーコードの検証と変換
            color_value = color_value.upper().strip()
            # 英数字のみを抽出
            color_value = ''.join(c for c in color_value if c in '0123456789ABCDEF')

            if len(color_value) == 6:
                # RGB形式の場合、先頭に'00'（アルファチャンネル）を追加
                color_value = '00' + color_value
            elif len(color_value) == 8:
                # すでにaRGB形式
                pass
            else:
                # 不正な形式の場合は色を設定しない
                color_value = None

            if color_value and len(color_value) == 8:
                normal_font_kwargs['color'] = color_value
        except Exception:
            # 色の変換に失敗した場合はスキップ
            pass

    # ハイライト色
    blue_font_kwargs['color'] = highlight_color

    # 下線
    if original_font.underline:
        normal_font_kwargs['u'] = original_font.underline
        blue_font_kwargs['u'] = original_font.underline

    # フォントオブジェクトを作成
    normal_font = InlineFont(**{k: v for k, v in normal_font_kwargs.items() if v is not None})
    blue_font = InlineFont(**{k: v for k, v in blue_font_kwargs.items() if v is not None})

    for start, end in differences:
        # 差分の前の通常テキスト
        if current_pos < start:
            rich_text_parts.append(TextBlock(normal_font, new_text[current_pos:start]))

        # 差分部分（青色）
        if start < end:
            rich_text_parts.append(TextBlock(blue_font, new_text[start:end]))

        current_pos = end

    # 残りのテキスト
    if current_pos < len(new_text):
        rich_text_parts.append(TextBlock(normal_font, new_text[current_pos:]))

    # セルにRichTextを設定
    if rich_text_parts:
        cell.value = CellRichText(*rich_text_parts)
    
    return diff_type


def compare_and_highlight_excel(old_file_path: str, new_file_path: str, output_file_path: str, highlight_color: str = DEFAULT_HIGHLIGHT_COLOR, compare_formulas: bool = False) -> List[Dict]:
    """
    2つのExcelファイルを比較し、差分を指定色でハイライト

    Args:
        old_file_path: 旧ファイルのパス
        new_file_path: 新ファイルのパス
        output_file_path: 出力ファイルのパス
        highlight_color: ハイライト色（aRGB形式の16進数、デフォルトは青）
        compare_formulas: Trueの場合は数式を比較、Falseの場合は表示値を比較
    
    Returns:
        changes_log: 変更履歴のリスト
    """
    print(f"\n処理開始...")
    print(f"古いファイル: {old_file_path}")
    print(f"新しいファイル: {new_file_path}")
    print(f"比較モード: {'数式' if compare_formulas else '表示値'}")

    # 処理開始時刻を記録
    start_time = time.time()

    # ファイルを開く（compare_formulasがTrueの場合は数式を保持、Falseの場合は表示値のみ）
    try:
        old_wb = openpyxl.load_workbook(old_file_path, data_only=not compare_formulas)
    except Exception as e:
        print(f"エラー: 旧ファイルを開けませんでした: {e}")
        raise

    try:
        new_wb = openpyxl.load_workbook(new_file_path, data_only=not compare_formulas)
    except Exception as e:
        print(f"エラー: 新ファイルを開けませんでした: {e}")
        old_wb.close()
        raise

    changes_log = []  # 変更履歴を記録

    # 全シートを比較
    for sheet_name in new_wb.sheetnames:
        if sheet_name not in old_wb.sheetnames:
            print(f"警告: シート '{sheet_name}' は古いファイルに存在しません")
            continue

        old_sheet = old_wb[sheet_name]
        new_sheet = new_wb[sheet_name]

        print(f"\nシート '{sheet_name}' を処理中...")
        sheet_changes = 0

        # 総セル数を計算
        total_cells = new_sheet.max_row * new_sheet.max_column
        processed_cells = 0
        last_progress = 0

        # 各セルを比較
        for row in range(1, new_sheet.max_row + 1):
            for col in range(1, new_sheet.max_column + 1):
                old_cell = old_sheet.cell(row, col)
                new_cell = new_sheet.cell(row, col)
                
                # 結合セルの処理
                from openpyxl.cell.cell import MergedCell
                if isinstance(new_cell, MergedCell):
                    # 結合セルはスキップ（マスターセルのみ処理される）
                    processed_cells += 1
                    continue

                old_value = get_cell_value_as_string(old_cell)
                new_value = get_cell_value_as_string(new_cell)

                # 両方空セルの場合はスキップ（パフォーマンス向上）
                if not old_value and not new_value:
                    processed_cells += 1
                    continue

                # 差分がある場合
                if old_value != new_value:
                    diff_type = 'equal'
                    if new_value:
                        diff_type = apply_blue_color_to_differences(new_cell, old_value, new_value, highlight_color)
                    elif old_value:
                        # 新値が空の場合は削除
                        diff_type = 'delete'
                    
                    sheet_changes += 1

                    # 変更履歴を記録
                    changes_log.append({
                        'sheet': sheet_name,
                        'cell': f'{new_cell.column_letter}{new_cell.row}',
                        'old': old_value[:MAX_CELL_VALUE_LENGTH] + ('...' if len(old_value) > MAX_CELL_VALUE_LENGTH else ''),
                        'new': new_value[:MAX_CELL_VALUE_LENGTH] + ('...' if len(new_value) > MAX_CELL_VALUE_LENGTH else ''),
                        'type': diff_type
                    })

                # 進行状況を表示（10%刻み）
                processed_cells += 1
                progress = int((processed_cells / total_cells) * 100)
                if progress >= last_progress + PROGRESS_DISPLAY_INTERVAL and progress < 100:
                    print(f"  進行状況: {progress}% ({processed_cells}/{total_cells} セル)")
                    last_progress = progress
    # 差分サマリーシートを作成
    if changes_log:
        print(f"\n差分サマリーシートを作成中...")
        summary_sheet = new_wb.create_sheet(SUMMARY_SHEET_NAME, 0)  # 最初のシートとして追加

        # ヘッダー行を追加
        summary_sheet['A1'] = 'No.'
        summary_sheet['B1'] = 'シート名'
        summary_sheet['C1'] = 'セル'
        summary_sheet['D1'] = '旧値'
        summary_sheet['E1'] = '新値'

        # ヘッダーのスタイル設定
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, size=SUMMARY_HEADER_FONT_SIZE)
        header_fill = PatternFill(start_color=SUMMARY_HEADER_COLOR, end_color=SUMMARY_HEADER_COLOR, fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in summary_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 変更履歴を書き込み
        for idx, change in enumerate(changes_log, start=2):
            summary_sheet[f'A{idx}'] = idx - 1
            summary_sheet[f'B{idx}'] = change['sheet']
            summary_sheet[f'C{idx}'] = change['cell']
            summary_sheet[f'D{idx}'] = change['old']
            summary_sheet[f'E{idx}'] = change['new']

        # 列幅を調整
        summary_sheet.column_dimensions['A'].width = SUMMARY_COL_WIDTH_NO
        summary_sheet.column_dimensions['B'].width = SUMMARY_COL_WIDTH_SHEET
        summary_sheet.column_dimensions['C'].width = SUMMARY_COL_WIDTH_CELL
        summary_sheet.column_dimensions['D'].width = SUMMARY_COL_WIDTH_VALUE
        summary_sheet.column_dimensions['E'].width = SUMMARY_COL_WIDTH_VALUE

        print(f"  差分サマリーシートに {len(changes_log)} 件の変更を記録しました")

    # 結果を保存
    new_wb.save(output_file_path)

    # 処理時間を計算
    elapsed_time = time.time() - start_time

    # 結果表示
    if len(changes_log) == 0:
        print(f"\n完了！ 差分は見つかりませんでした（新旧ファイルは同一です）")
    else:
        print(f"\n完了！ 合計 {len(changes_log)} 個のセルに差分が見つかりました")
    print(f"処理時間: {elapsed_time:.1f}秒")
    print(f"出力ファイル: {output_file_path}")

    old_wb.close()
    new_wb.close()
    
    return changes_log


def generate_html_report(all_results: List[Dict], output_path: str, color_name: str, mode_name: str, total_time: float):
    """
    差分結果からHTMLレポートを生成

    Args:
        all_results: 全ファイルの差分結果リスト
        output_path: 出力先パス
        color_name: 使用したハイライト色名
        mode_name: 比較モード名
        total_time: 総処理時間
    """
    # 統計情報を計算
    total_files = len(all_results)
    total_changes = sum(len(result['changes']) for result in all_results)
    success_files = sum(1 for result in all_results if result['status'] == 'success')
    error_files = total_files - success_files
    
    # シート別統計
    sheet_stats = {}
    for result in all_results:
        for change in result['changes']:
            sheet_name = change['sheet']
            if sheet_name not in sheet_stats:
                sheet_stats[sheet_name] = 0
            sheet_stats[sheet_name] += 1
    
    # ファイル別統計（グラフ用）
    file_stats = [(result['base_name'], len(result['changes'])) for result in all_results]
    
    # 現在時刻
    generated_time = datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')
    
    # HTMLテンプレート
    html_content = f"""<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel差分レポート</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.0/font/bootstrap-icons.css" rel="stylesheet">
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
    <style>
        :root {{
            --bg-primary: #ffffff;
            --bg-secondary: #f8f9fa;
            --text-primary: #212529;
            --text-secondary: #6c757d;
            --border-color: #dee2e6;
            --diff-old: #ffe5e5;
            --diff-new: #e5f5ff;
        }}
        
        [data-bs-theme="dark"] {{
            --bg-primary: #1a1d20;
            --bg-secondary: #2b3035;
            --text-primary: #e9ecef;
            --text-secondary: #adb5bd;
            --border-color: #495057;
            --diff-old: #4a2020;
            --diff-new: #1a3a4a;
        }}
        
        body {{
            background-color: var(--bg-secondary);
            color: var(--text-primary);
            transition: background-color 0.3s, color 0.3s;
        }}
        
        .card {{
            background-color: var(--bg-primary);
            border-color: var(--border-color);
            margin-bottom: 1.5rem;
        }}
        
        .stat-card {{
            border-left: 4px solid #0d6efd;
        }}
        
        .stat-card.success {{
            border-left-color: #198754;
        }}
        
        .stat-card.warning {{
            border-left-color: #ffc107;
        }}
        
        .stat-card.danger {{
            border-left-color: #dc3545;
        }}
        
        .file-accordion .accordion-button {{
            background-color: var(--bg-secondary);
            color: var(--text-primary);
            display: flex;
            align-items: center;
            gap: 0.5rem;
        }}
        
        .file-accordion .accordion-button:not(.collapsed) {{
            background-color: #0d6efd;
            color: white;
        }}
        
        .file-name-text {{
            flex: 1;
            min-width: 0;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }}
        
        .badge-group {{
            display: flex;
            gap: 0.25rem;
            margin-left: auto;
            flex-shrink: 0;
        }}
        
        .diff-table {{
            font-size: 0.9rem;
        }}
        
        .diff-old {{
            background-color: var(--diff-old);
        }}
        
        .diff-new {{
            background-color: var(--diff-new);
        }}
        
        .badge-custom {{
            font-size: 0.75rem;
            padding: 0.35em 0.65em;
        }}
        
        .badge-insert {{
            background-color: #198754;
            color: #ffffff;
            font-weight: bold;
            padding: 0.35em 0.65em;
            border-radius: 0.25rem;
        }}
        
        .badge-delete {{
            background-color: #dc3545;
            color: #ffffff;
            font-weight: bold;
            padding: 0.35em 0.65em;
            border-radius: 0.25rem;
        }}
        
        .badge-replace {{
            background-color: #0d6efd;
            color: #ffffff;
            font-weight: bold;
            padding: 0.35em 0.65em;
            border-radius: 0.25rem;
        }}
        
        .search-highlight {{
            background-color: yellow;
            color: black;
            font-weight: bold;
        }}
        
        .filter-section {{
            position: sticky;
            top: 0;
            z-index: 1000;
            background-color: var(--bg-primary);
            padding: 1rem;
            border-bottom: 2px solid var(--border-color);
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        
        .chart-container {{
            position: relative;
            height: 300px;
        }}
        
        @media print {{
            .filter-section, .no-print {{
                display: none;
            }}
        }}
    </style>
</head>
<body>
    <nav class="navbar navbar-expand-lg navbar-dark bg-primary mb-4">
        <div class="container-fluid">
            <span class="navbar-brand mb-0 h1">
                <i class="bi bi-file-earmark-diff"></i> Excel差分レポート
            </span>
            <div class="d-flex">
                <button class="btn btn-outline-light me-2" onclick="toggleDarkMode()">
                    <i class="bi bi-moon-stars"></i>
                </button>
                <button class="btn btn-outline-light" onclick="window.print()">
                    <i class="bi bi-printer"></i> 印刷
                </button>
            </div>
        </div>
    </nav>

    <div class="container-fluid">
        <!-- サマリーセクション -->
        <div class="row mb-4">
            <div class="col-12">
                <div class="card">
                    <div class="card-body">
                        <h5 class="card-title">
                            <i class="bi bi-info-circle"></i> 処理サマリー
                        </h5>
                        <p class="text-muted mb-3">
                            生成日時: {generated_time}<br>
                            ハイライト色: {html.escape(color_name)} | 比較モード: {html.escape(mode_name)}
                        </p>
                        <div class="row">
                            <div class="col-md-3">
                                <div class="card stat-card">
                                    <div class="card-body text-center">
                                        <h3 class="mb-0">{total_files}</h3>
                                        <small class="text-muted">処理ファイル数</small>
                                    </div>
                                </div>
                            </div>
                            <div class="col-md-3">
                                <div class="card stat-card warning">
                                    <div class="card-body text-center">
                                        <h3 class="mb-0">{total_changes}</h3>
                                        <small class="text-muted">総差分数</small>
                                    </div>
                                </div>
                            </div>
                            <div class="col-md-3">
                                <div class="card stat-card success">
                                    <div class="card-body text-center">
                                        <h3 class="mb-0">{success_files}</h3>
                                        <small class="text-muted">成功</small>
                                    </div>
                                </div>
                            </div>
                            <div class="col-md-3">
                                <div class="card stat-card {'danger' if error_files > 0 else ''}">
                                    <div class="card-body text-center">
                                        <h3 class="mb-0">{error_files}</h3>
                                        <small class="text-muted">エラー</small>
                                    </div>
                                </div>
                            </div>
                        </div>
                        <div class="mt-3 text-center">
                            <small class="text-muted">
                                <i class="bi bi-clock"></i> 処理時間: {total_time:.1f}秒
                            </small>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- グラフセクション -->
        <div class="row mb-4">
            <div class="col-md-6">
                <div class="card">
                    <div class="card-body">
                        <h6 class="card-title">ファイル別差分数</h6>
                        <div class="chart-container">
                            <canvas id="fileChart"></canvas>
                        </div>
                    </div>
                </div>
            </div>
            <div class="col-md-6">
                <div class="card">
                    <div class="card-body">
                        <h6 class="card-title">シート別差分数</h6>
                        <div class="chart-container">
                            <canvas id="sheetChart"></canvas>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- フィルタ・検索セクション -->
        <div class="filter-section mb-3 no-print">
            <div class="row g-3">
                <div class="col-md-3">
                    <input type="text" class="form-control" id="searchInput" 
                           placeholder="🔍 差分内容を検索...">
                </div>
                <div class="col-md-2">
                    <select class="form-select" id="sortSelect">
                        <option value="name-asc">ファイル名 (昇順)</option>
                        <option value="name-desc">ファイル名 (降順)</option>
                        <option value="diff-desc">差分数 (多い順)</option>
                        <option value="diff-asc">差分数 (少ない順)</option>
                        <option value="original">元の順序</option>
                    </select>
                </div>
                <div class="col-md-2">
                    <select class="form-select" id="fileFilter">
                        <option value="">すべてのファイル</option>
                        {generate_file_filter_options(all_results)}
                    </select>
                </div>
                <div class="col-md-3">
                    <select class="form-select" id="sheetFilter">
                        <option value="">すべてのシート</option>
                        {generate_sheet_filter_options(all_results)}
                    </select>
                </div>
                <div class="col-md-2">
                    <button class="btn btn-secondary w-100" onclick="resetFilters()">
                        <i class="bi bi-arrow-counterclockwise"></i> リセット
                    </button>
                </div>
            </div>
        </div>

        <!-- 差分詳細セクション -->
        <div class="accordion file-accordion" id="diffAccordion">
            {generate_accordion_items(all_results)}
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
    <script>
        // ダークモード切り替え
        function toggleDarkMode() {{
            const html = document.documentElement;
            const currentTheme = html.getAttribute('data-bs-theme');
            html.setAttribute('data-bs-theme', currentTheme === 'dark' ? 'light' : 'dark');
            localStorage.setItem('theme', currentTheme === 'dark' ? 'light' : 'dark');
            updateCharts();
        }}
        
        // テーマの復元
        const savedTheme = localStorage.getItem('theme') || 'light';
        document.documentElement.setAttribute('data-bs-theme', savedTheme);
        
        // グラフデータ
        const fileData = {json.dumps(file_stats)};
        const sheetData = {json.dumps(list(sheet_stats.items()))};
        
        let fileChart, sheetChart;
        
        function getChartColors() {{
            const isDark = document.documentElement.getAttribute('data-bs-theme') === 'dark';
            return {{
                textColor: isDark ? '#e9ecef' : '#212529',
                gridColor: isDark ? '#495057' : '#dee2e6'
            }};
        }}
        
        function updateCharts() {{
            const colors = getChartColors();
            
            if (fileChart) {{
                fileChart.options.scales.y.ticks.color = colors.textColor;
                fileChart.options.scales.y.grid.color = colors.gridColor;
                fileChart.options.scales.x.ticks.color = colors.textColor;
                fileChart.options.scales.x.grid.color = colors.gridColor;
                fileChart.options.plugins.legend.labels.color = colors.textColor;
                fileChart.update();
            }}
            
            if (sheetChart) {{
                sheetChart.options.plugins.legend.labels.color = colors.textColor;
                sheetChart.update();
            }}
        }}
        
        // ファイル別グラフ
        const fileCtx = document.getElementById('fileChart').getContext('2d');
        fileChart = new Chart(fileCtx, {{
            type: 'bar',
            data: {{
                labels: fileData.map(d => d[0]),
                datasets: [{{
                    label: '差分数',
                    data: fileData.map(d => d[1]),
                    backgroundColor: 'rgba(13, 110, 253, 0.5)',
                    borderColor: 'rgba(13, 110, 253, 1)',
                    borderWidth: 1
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                scales: {{
                    y: {{
                        beginAtZero: true,
                        ticks: {{ color: getChartColors().textColor }},
                        grid: {{ color: getChartColors().gridColor }}
                    }},
                    x: {{
                        ticks: {{ 
                            color: getChartColors().textColor,
                            maxRotation: 45,
                            minRotation: 45
                        }},
                        grid: {{ color: getChartColors().gridColor }}
                    }}
                }},
                plugins: {{
                    legend: {{
                        labels: {{ color: getChartColors().textColor }}
                    }}
                }}
            }}
        }});
        
        // シート別グラフ
        const sheetCtx = document.getElementById('sheetChart').getContext('2d');
        sheetChart = new Chart(sheetCtx, {{
            type: 'doughnut',
            data: {{
                labels: sheetData.map(d => d[0]),
                datasets: [{{
                    data: sheetData.map(d => d[1]),
                    backgroundColor: [
                        'rgba(13, 110, 253, 0.7)',
                        'rgba(25, 135, 84, 0.7)',
                        'rgba(255, 193, 7, 0.7)',
                        'rgba(220, 53, 69, 0.7)',
                        'rgba(108, 117, 125, 0.7)',
                        'rgba(13, 202, 240, 0.7)'
                    ]
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                plugins: {{
                    legend: {{
                        position: 'right',
                        labels: {{ color: getChartColors().textColor }}
                    }}
                }}
            }}
        }});
        
        // 検索機能
        document.getElementById('searchInput').addEventListener('input', function() {{
            const searchText = this.value.toLowerCase();
            filterResults();
        }});
        
        // フィルタ機能
        document.getElementById('fileFilter').addEventListener('change', filterResults);
        document.getElementById('sheetFilter').addEventListener('change', filterResults);
        
        // ソート機能
        document.getElementById('sortSelect').addEventListener('change', function() {{
            sortAccordionItems(this.value);
        }});
        
        function sortAccordionItems(sortType) {{
            const accordion = document.getElementById('diffAccordion');
            const items = Array.from(accordion.querySelectorAll('.accordion-item'));
            
            items.sort((a, b) => {{
                const nameA = a.dataset.fileName || '';
                const nameB = b.dataset.fileName || '';
                const diffA = parseInt(a.dataset.diffCount) || 0;
                const diffB = parseInt(b.dataset.diffCount) || 0;
                const orderA = parseInt(a.dataset.originalOrder) || 0;
                const orderB = parseInt(b.dataset.originalOrder) || 0;
                
                switch(sortType) {{
                    case 'name-asc':
                        return nameA.localeCompare(nameB, 'ja');
                    case 'name-desc':
                        return nameB.localeCompare(nameA, 'ja');
                    case 'diff-desc':
                        return diffB - diffA;
                    case 'diff-asc':
                        return diffA - diffB;
                    case 'original':
                        return orderA - orderB;
                    default:
                        return 0;
                }}
            }});
            
            // アコーディオンを再構築
            items.forEach(item => accordion.appendChild(item));
        }}
        
        function filterResults() {{
            const searchText = document.getElementById('searchInput').value.toLowerCase();
            const selectedFile = document.getElementById('fileFilter').value;
            const selectedSheet = document.getElementById('sheetFilter').value;
            
            document.querySelectorAll('.accordion-item').forEach(item => {{
                const fileName = item.dataset.fileName;
                let visible = true;
                
                // ファイルフィルタ
                if (selectedFile && fileName !== selectedFile) {{
                    visible = false;
                }}
                
                // シート・検索フィルタ
                if (visible && (selectedSheet || searchText)) {{
                    const rows = item.querySelectorAll('tbody tr');
                    let hasVisibleRow = false;
                    
                    rows.forEach(row => {{
                        const sheetName = row.dataset.sheet;
                        const oldValue = row.cells[3].textContent.toLowerCase();
                        const newValue = row.cells[4].textContent.toLowerCase();
                        
                        let rowVisible = true;
                        
                        if (selectedSheet && sheetName !== selectedSheet) {{
                            rowVisible = false;
                        }}
                        
                        if (searchText && !oldValue.includes(searchText) && !newValue.includes(searchText)) {{
                            rowVisible = false;
                        }}
                        
                        row.style.display = rowVisible ? '' : 'none';
                        if (rowVisible) hasVisibleRow = true;
                    }});
                    
                    visible = hasVisibleRow;
                }}
                
                item.style.display = visible ? '' : 'none';
            }});
        }}
        
        function resetFilters() {{
            document.getElementById('searchInput').value = '';
            document.getElementById('fileFilter').value = '';
            document.getElementById('sheetFilter').value = '';
            document.getElementById('sortSelect').value = 'name-asc';
            sortAccordionItems('name-asc');
            filterResults();
        }}
    </script>
</body>
</html>"""
    
    # HTMLファイルを保存
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"\nHTMLレポートを生成しました: {output_path}")


def generate_file_filter_options(all_results: List[Dict]) -> str:
    """ファイルフィルタのオプションを生成"""
    options = []
    for result in all_results:
        if result['changes']:
            name = html.escape(result['base_name'])
            options.append(f'<option value="{name}">{name}</option>')
    return '\n'.join(options)


def generate_sheet_filter_options(all_results: List[Dict]) -> str:
    """シートフィルタのオプションを生成"""
    sheets = set()
    for result in all_results:
        for change in result['changes']:
            sheets.add(change['sheet'])
    
    options = []
    for sheet in sorted(sheets):
        name = html.escape(sheet)
        options.append(f'<option value="{name}">{name}</option>')
    return '\n'.join(options)


def generate_accordion_items(all_results: List[Dict]) -> str:
    """差分詳細のアコーディオンアイテムを生成"""
    items = []
    
    for i, result in enumerate(all_results):
        file_name = html.escape(result['base_name'])
        changes = result['changes']
        change_count = len(changes)
        status = result.get('status', 'success')
        
        # 種類別カウント
        insert_count = sum(1 for c in changes if c.get('type') == 'insert')
        delete_count = sum(1 for c in changes if c.get('type') == 'delete')
        replace_count = sum(1 for c in changes if c.get('type') == 'replace')
        
        # ステータスに応じてバッジとアイコンを設定
        if status == 'error':
            icon = 'x-circle'
            badge_html = '<span class="badge bg-danger">エラー</span>'
        elif change_count == 0:
            icon = 'check-circle'
            badge_html = '<span class="badge bg-success">差分なし</span>'
        else:
            icon = 'exclamation-triangle'
            # 種類別バッジを生成
            badges = []
            badges.append(f'<span class="badge bg-warning">{change_count}件</span>')
            if insert_count > 0:
                badges.append(f'<span class="badge badge-insert">{insert_count}追加</span>')
            if delete_count > 0:
                badges.append(f'<span class="badge badge-delete">{delete_count}削除</span>')
            if replace_count > 0:
                badges.append(f'<span class="badge badge-replace">{replace_count}変更</span>')
            badge_html = ''.join(badges)
        
        # エラーの場合はエラー情報を表示
        if status == 'error':
            error_message = html.escape(result.get('error', '不明なエラー'))
            tables_html = f'''
                <tr>
                    <td colspan="6">
                        <div class="alert alert-danger" role="alert">
                            <h6 class="alert-heading">
                                <i class="bi bi-exclamation-octagon"></i> エラーが発生しました
                            </h6>
                            <hr>
                            <p class="mb-0"><strong>エラー内容:</strong></p>
                            <pre class="mt-2 mb-0" style="background-color: #f8d7da; padding: 1rem; border-radius: 0.25rem; font-size: 0.85rem; color: #842029;">{error_message}</pre>
                            <hr>
                            <small class="text-muted">
                                <strong>旧ファイル:</strong> {html.escape(result.get('old_file', 'N/A'))}<br>
                                <strong>新ファイル:</strong> {html.escape(result.get('new_file', 'N/A'))}
                            </small>
                        </div>
                    </td>
                </tr>
            '''
        else:
            # テーブル行を生成
            table_rows = []
            for idx, change in enumerate(changes, 1):
                sheet = html.escape(change['sheet'])
                cell = html.escape(change['cell'])
                old_val = html.escape(change['old'])
                new_val = html.escape(change['new'])
                diff_type = change.get('type', 'replace')
                
                # 差分タイプに応じたバッジとクラス
                if diff_type == 'insert':
                    type_badge = '<span class="badge badge-insert">追加</span>'
                    row_class = 'diff-type-insert'
                elif diff_type == 'delete':
                    type_badge = '<span class="badge badge-delete">削除</span>'
                    row_class = 'diff-type-delete'
                else:
                    type_badge = '<span class="badge badge-replace">変更</span>'
                    row_class = 'diff-type-replace'
                
                table_rows.append(f'''
                    <tr data-sheet="{sheet}" class="{row_class}">
                        <td>{idx}</td>
                        <td><span class="badge bg-secondary">{sheet}</span></td>
                        <td><code>{cell}</code></td>
                        <td class="diff-old">{old_val}</td>
                        <td class="diff-new">{new_val}</td>
                        <td>{type_badge}</td>
                    </tr>
                ''')
            
            tables_html = '\n'.join(table_rows) if table_rows else '<tr><td colspan="6" class="text-center text-muted">差分なし</td></tr>'
        
        item_html = f'''
            <div class="accordion-item" data-file-name="{file_name}" data-diff-count="{change_count}" data-original-order="{i}">
                <h2 class="accordion-header">
                    <button class="accordion-button collapsed" type="button" 
                            data-bs-toggle="collapse" data-bs-target="#collapse{i}">
                        <i class="bi bi-{icon} me-2"></i>
                        <span class="file-name-text">{file_name}</span>
                        <span class="badge-group">{badge_html}</span>
                    </button>
                </h2>
                <div id="collapse{i}" class="accordion-collapse collapse" 
                     data-bs-parent="#diffAccordion">
                    <div class="accordion-body">
                        <div class="table-responsive">
                            <table class="table table-sm table-hover diff-table">
                                <thead class="table-light">
                                    <tr>
                                        <th style="width: 5%">No.</th>
                                        <th style="width: 12%">シート</th>
                                        <th style="width: 8%">セル</th>
                                        <th style="width: 30%">旧値</th>
                                        <th style="width: 30%">新値</th>
                                        <th style="width: 10%">種類</th>
                                    </tr>
                                </thead>
                                <tbody>
                                    {tables_html}
                                </tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>
        '''
        items.append(item_html)
    
    return '\n'.join(items)


def main():
    """
    メイン処理
    """
    print("=" * SEPARATOR_LENGTH)
    print("Excel差分ハイライトツール")
    print("=" * SEPARATOR_LENGTH)
    print("\n旧バージョンと新バージョンがそれぞれ別のフォルダにあり、")
    print("同じファイル名（ベース名）のペアをすべて自動処理します。")

    # ハイライト色の選択
    print("\n差分をハイライトする色を選択してください:")
    for key, (_, color_name) in sorted(COLOR_MAP.items()):
        print(f"{key}. {color_name}")

    color_choice = input(f"選択 (1-{len(COLOR_MAP)}, デフォルト: 1): ").strip()

    if color_choice not in COLOR_MAP:
        color_choice = '1'  # デフォルトは青

    highlight_color, color_name = COLOR_MAP[color_choice]
    print(f"選択された色: {color_name}\n")

    # 比較モードの選択
    print("差分比較のモードを選択してください:")
    print("1. 表示値のみ比較（数式は比較しない）")
    print("2. 数式を比較（数式がある場合は数式を比較）")

    mode_choice = input("選択 (1-2, デフォルト: 1): ").strip()
    compare_formulas = (mode_choice == '2')

    mode_name = "数式" if compare_formulas else "表示値"
    print(f"選択されたモード: {mode_name}\n")

    # フォルダ配下の全ファイルを一括処理
    print("旧バージョンのフォルダを指定してください")
    old_directory = input("旧バージョンのフォルダパス: ").strip().strip('"').strip("'")
    if not old_directory:
        old_directory = "."

    print("\n新バージョンのフォルダを指定してください")
    new_directory = input("新バージョンのフォルダパス: ").strip().strip('"').strip("'")
    if not new_directory:
        new_directory = "."

    print("\n出力先フォルダを指定してください")
    output_directory = input("出力先フォルダパス（空欄で新バージョンと同じ）: ").strip().strip('"').strip("'")
    if not output_directory:
        output_directory = new_directory

    # 処理開始時刻を記録（全体の処理時間計測用）
    main_start_time = time.time()

    # マッチングするファイルペアを検索
    file_pairs, unmatched_old_files, unmatched_new_files = find_matching_file_pairs(old_directory, new_directory)

    if not file_pairs:
        print("\nマッチングするファイルペアが見つかりませんでした")
        return

    print(f"\n{len(file_pairs)} 個のファイルペアが見つかりました:")
    for i, (base_name, old_file, new_file) in enumerate(file_pairs, 1):
        print(f"{i}. {base_name}")
        print(f"   旧: {Path(old_file).name}")
        print(f"   新: {Path(new_file).name}")

    if not file_pairs:
        print("\nファイルが見つからなかったため処理を終了します")
        return

    # 出力ディレクトリを作成（存在しない場合）
    output_path = Path(output_directory)
    if not output_path.exists():
        output_path.mkdir(parents=True, exist_ok=True)

    # 確認
    if len(file_pairs) == 1:
        base_name, old_file, new_file = file_pairs[0]
        new_file_path = Path(new_file)
        output_filename = new_file_path.stem + OUTPUT_FILE_SUFFIX + new_file_path.suffix
        output_file = str(output_path / output_filename)
        print(f"\n出力ファイル: {output_filename}")
    else:
        print(f"\n出力先: {output_directory}")
        print(f"処理対象: {len(file_pairs)} ファイル")

    confirm = input("\n処理を開始しますか？ (y/n): ").strip().lower()

    if confirm != 'y':
        print("処理をキャンセルしました")
        return

    # 比較とハイライト処理
    success_count = 0
    error_count = 0
    all_results = []  # 全ファイルの結果を記録

    for i, (base_name, old_file, new_file) in enumerate(file_pairs, 1):
        try:
            print(f"\n{'='*SEPARATOR_LENGTH}")
            print(f"[{i}/{len(file_pairs)}] 処理中: {Path(new_file).name}")
            print(f"{'='*SEPARATOR_LENGTH}")

            new_file_path = Path(new_file)
            output_filename = new_file_path.stem + OUTPUT_FILE_SUFFIX + new_file_path.suffix
            output_file = str(output_path / output_filename)

            changes = compare_and_highlight_excel(old_file, new_file, output_file, highlight_color, compare_formulas)
            success_count += 1
            
            # 結果を記録
            all_results.append({
                'base_name': base_name,
                'old_file': Path(old_file).name,
                'new_file': Path(new_file).name,
                'output_file': output_filename,
                'changes': changes,
                'status': 'success'
            })

        except Exception as e:
            print(f"\nエラーが発生しました: {e}")
            error_count += 1
            all_results.append({
                'base_name': base_name,
                'old_file': Path(old_file).name if old_file else 'N/A',
                'new_file': Path(new_file).name if new_file else 'N/A',
                'output_file': 'N/A',
                'changes': [],
                'status': 'error',
                'error': str(e)
            })
            import traceback
            traceback.print_exc()

    # 最終結果
    total_time = time.time() - main_start_time  # 全体の処理時間を計算
    
    print(f"\n{'='*SEPARATOR_LENGTH}")
    print(f"処理完了")
    print(f"{'='*SEPARATOR_LENGTH}")
    print(f"成功: {success_count} ファイル")
    print(f"失敗: {error_count} ファイル")
    print(f"出力先: {output_directory}")

    # HTMLレポートを生成
    if all_results:
        html_filename = f"diff_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        html_path = str(output_path / html_filename)
        generate_html_report(all_results, html_path, color_name, mode_name, total_time)

    # マッチングしなかったファイルの報告
    if unmatched_old_files:
        print(f"\n旧フォルダにのみ存在（{len(unmatched_old_files)} ファイル）:")
        for filename in unmatched_old_files:
            print(f"  - {filename}")

    if unmatched_new_files:
        print(f"\n新フォルダにのみ存在（{len(unmatched_new_files)} ファイル）:")
        for filename in unmatched_new_files:
            print(f"  - {filename}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n処理を中断しました")
    except Exception as e:
        print(f"\n\n予期しないエラーが発生しました: {e}")
        import traceback
        traceback.print_exc()
    finally:
        input("\nEnterキーを押して終了...")
